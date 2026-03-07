#!/usr/bin/env python3
"""
Research index builder + lightweight RAG router.

What it does
------------
1) Walks a research folder (or uses a provenance manifest).
2) Deduplicates files by SHA-256.
3) Extracts lightweight metadata/text from PDFs, text/markdown, RIS, XLSX, and images.
4) Builds a JSON index with:
   - summaries/findings
   - metadata
   - file locations
   - topic tags
   - routing terms / routing text
5) Routes queries to the most relevant files using a lightweight lexical scorer.

This is intentionally local-first and dependency-light.
Optional packages:
  pip install pypdf openpyxl pillow

Examples
--------
Build an index:
    python research_index_router.py build \
        --root /path/to/External_Research \
        --output external_research_index.json \
        --manifest /path/to/RESEARCH_MANIFEST.json

Route a query:
    python research_index_router.py route \
        --index external_research_index.json \
        --query "quantum reaction diffusion universality and decay exponents" \
        --top-k 8

Print routing results as pretty JSON:
    python research_index_router.py route \
        --index external_research_index.json \
        --query "metriplectic dissipative Hamiltonian systems" \
        --top-k 5 --pretty
"""
from __future__ import annotations

import argparse
import json
import math
import mimetypes
import os
import re
import sys
import hashlib
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

try:
    from pypdf import PdfReader  # type: ignore
except Exception:  # pragma: no cover
    PdfReader = None

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None

try:
    from PIL import Image  # type: ignore
except Exception:  # pragma: no cover
    Image = None


TEXT_EXTS = {
    ".txt", ".md", ".markdown", ".rst", ".tex", ".csv", ".json", ".yaml", ".yml",
    ".py", ".ipynb", ".toml", ".ini", ".cfg"
}
IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".tif", ".tiff"}
STOPWORDS = {
    "a","an","and","are","as","at","be","been","but","by","can","do","does","for","from",
    "has","have","if","in","into","is","it","its","of","on","or","s","such","that","the",
    "their","this","to","using","via","we","with","within","without","what","which","who",
    "why","how","than","then","these","those","our","your","they","them","he","she","you",
    "paper","study","studies","article","book","chapter","notes","note"
}
FINDING_VERBS = (
    "show", "find", "demonstrate", "derive", "predict", "propose", "present",
    "report", "introduce", "argue", "reveal", "establish", "analyze", "analyse",
    "develop", "construct", "explain", "confirm", "suggest"
)
DEFAULT_EXCLUDES = {
    "dirs": {
        ".cache", ".git", ".idea", ".mypy_cache", ".pytest_cache", ".venv", ".vscode",
        "__pycache__", "build", "dist", "node_modules", "out", "venv"
    },
    "file_names": {"PROVENANCE_manifest.json", ".DS_Store"},
    "file_suffixes": {".dll", ".dylib", ".exe", ".o", ".obj", ".pdb", ".pyc", ".pyo", ".so"}
}


@dataclass
class FileRecord:
    abs_path: Path
    rel_path: str
    sha256: str
    size: int
    modified_time_utc: str


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def file_mtime_iso(path: Path) -> str:
    ts = path.stat().st_mtime
    return datetime.fromtimestamp(ts, timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def safe_rel_path(path: Path, root: Path) -> str:
    return path.relative_to(root).as_posix()


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        while True:
            chunk = f.read(chunk_size)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def clean_text(text: str) -> str:
    text = text.replace("\x00", " ")
    text = text.replace("-\n", "")
    text = text.replace("\r", "\n")
    text = re.sub(r"\n+", " ", text)
    text = re.sub(r"[ \t]+", " ", text)
    return text.strip()


def compact(text: str, limit: int = 420) -> str:
    text = clean_text(text)
    if len(text) <= limit:
        return text
    cut = text[:limit]
    last = max(cut.rfind(". "), cut.rfind("; "), cut.rfind(", "))
    if last > limit * 0.55:
        cut = cut[:last + 1]
    return cut.rstrip(" .,;:") + "…"


def split_sentences(text: str) -> List[str]:
    text = clean_text(text)
    if not text:
        return []
    parts = re.split(r"(?<=[.!?])\s+(?=[A-Z0-9(])", text)
    return [p.strip() for p in parts if p.strip()]


def tokenize(text: str) -> List[str]:
    text = text.lower()
    text = text.replace("_", " ").replace("-", " ")
    tokens = re.findall(r"[a-z0-9]{2,}", text)
    return [t for t in tokens if t not in STOPWORDS]


def unique_keep_order(items: Iterable[str]) -> List[str]:
    seen = set()
    out = []
    for item in items:
        if not item:
            continue
        if item not in seen:
            seen.add(item)
            out.append(item)
    return out


def looks_like_title(value: Optional[str]) -> bool:
    if not value:
        return False
    value = clean_text(value)
    if len(value) < 6 or len(value) > 220:
        return False
    bad_prefixes = ("microsoft word", "untitled", "document", "default")
    low = value.lower()
    return not any(low.startswith(p) for p in bad_prefixes)


def filename_to_title(name: str) -> str:
    stem = Path(name).stem
    stem = re.sub(r"\b(copy|accepted|final|draft)\b", "", stem, flags=re.I)
    stem = stem.replace("_", " ").replace("-", " ")
    stem = re.sub(r"\s{2,}", " ", stem).strip()
    return stem or name


def infer_topics(rel_path: str) -> List[str]:
    parts = Path(rel_path).parts[:-1]
    drop = {"external_research", "external research", "rd", "research"}
    topics = []
    for part in parts:
        p = part.replace("_", " ").strip()
        if p.lower() in drop:
            continue
        if re.fullmatch(r"\d{4}\.\d{5}v\d+", p):
            continue
        topics.append(p)
    return unique_keep_order(topics)


def infer_doc_kind(ext: str, title: str, rel_path: str, snippet: str = "") -> str:
    low = f"{title} {rel_path} {snippet[:500]}".lower()
    if ext == ".ris":
        return "citation_export"
    if ext in IMAGE_EXTS:
        return "image"
    if ext == ".xlsx":
        return "dataset"
    if ext == ".pdf":
        if "supplementary" in low or "moesm" in low:
            return "supplementary_material"
        if "tutorial" in low or "introduction" in low or "monograph" in low:
            return "book_or_tutorial"
        if "annals" in low or "physrev" in low or "arxiv" in low or re.search(r"\bdoi\b", low):
            return "paper"
        return "paper"
    if ext in {".md", ".markdown"}:
        return "note"
    if ext == ".txt":
        if "references" in low or "bibliography" in low:
            return "reference_list"
        return "note_or_extract"
    return "document"


def guess_mime(path: Path) -> str:
    mime, _ = mimetypes.guess_type(str(path))
    return mime or "application/octet-stream"


def sentence_priority(sentence: str) -> int:
    low = sentence.lower()
    score = 0
    if any(v in low for v in FINDING_VERBS):
        score += 3
    if "we " in low or low.startswith("this "):
        score += 1
    if 40 <= len(sentence) <= 260:
        score += 1
    if re.search(r"\b(result|method|model|equation|framework|theory|experiment|analysis)\b", low):
        score += 1
    return score


def summarize_findings(title: str, text: str, doc_kind: str) -> Tuple[str, str]:
    text = clean_text(text)
    if not text:
        if doc_kind == "image":
            return ("Image asset; routing relies on filename and folder context rather than extracted text.", "filename_and_folder_context")
        if doc_kind == "citation_export":
            return ("Citation export file; useful for provenance and related-paper expansion rather than direct semantic retrieval.", "structured_metadata")
        if doc_kind == "dataset":
            return ("Spreadsheet/dataset asset; likely contains supplementary measurements or tabular evidence for nearby papers.", "sheet_metadata")
        return ("Limited extractable text; routing relies mostly on title, filename, and folder context.", "filename_and_folder_context")

    sents = split_sentences(text)
    if not sents:
        return (compact(text), "opening_text")

    ranked = sorted(
        enumerate(sents[:16]),
        key=lambda p: (sentence_priority(p[1]), -p[0]),
        reverse=True
    )
    chosen: List[str] = []
    for _, sent in ranked:
        if sent not in chosen:
            chosen.append(sent)
        if len(chosen) >= 2:
            break

    if not chosen:
        chosen = sents[:2]

    summary = " ".join(chosen)
    summary = compact(summary, 380)

    low = summary.lower()
    if doc_kind in {"paper", "book_or_tutorial"} and not any(v in low for v in FINDING_VERBS):
        lead = title if title else "This document"
        summary = compact(f"{lead}. {summary}", 380)

    basis = "abstract_or_opening_text"
    if re.search(r"\babstract\b", text[:1500], flags=re.I):
        basis = "abstract_or_opening_text"
    else:
        basis = "opening_text"
    return summary, basis


def extract_pdf(path: Path) -> Dict[str, Any]:
    meta: Dict[str, Any] = {
        "text": "",
        "title": filename_to_title(path.name),
        "page_count": None,
        "extraction_status": "unavailable",
        "notes": []
    }
    if PdfReader is None:
        meta["notes"].append("pypdf not installed")
        return meta

    try:
        reader = PdfReader(str(path))
        meta["page_count"] = len(reader.pages)
        doc_title = None
        try:
            raw_title = getattr(reader.metadata, "title", None) if reader.metadata else None
            if looks_like_title(raw_title):
                doc_title = clean_text(str(raw_title))
        except Exception:
            pass

        pages_text = []
        max_pages = min(6, len(reader.pages))
        for i in range(max_pages):
            try:
                pages_text.append(reader.pages[i].extract_text() or "")
            except Exception as e:
                meta["notes"].append(f"page_{i+1}_extract_error: {type(e).__name__}")
        joined = clean_text("\n".join(pages_text))

        if not doc_title and joined:
            lines = [ln.strip() for ln in joined.splitlines() if ln.strip()]
            for line in lines[:12]:
                if looks_like_title(line):
                    doc_title = line
                    break

        meta["title"] = doc_title or meta["title"]

        abstract = ""
        m = re.search(r"\babstract\b[:\s]*", joined, flags=re.I)
        if m:
            start = m.end()
            chunk = joined[start:start + 2200]
            stop = re.search(r"\b(keywords|introduction|1\.|i\.)\b", chunk, flags=re.I)
            abstract = chunk[:stop.start()] if stop else chunk

        meta["text"] = clean_text(abstract or joined[:4000])
        meta["extraction_status"] = "ok" if meta["text"] else "empty"
        return meta
    except Exception as e:
        meta["extraction_status"] = f"error:{type(e).__name__}"
        meta["notes"].append(str(e))
        return meta


def extract_text_file(path: Path) -> Dict[str, Any]:
    out = {"text": "", "title": filename_to_title(path.name), "line_count": None, "extraction_status": "ok"}
    try:
        text = path.read_text(encoding="utf-8", errors="ignore")
        text = clean_text(text)
        out["line_count"] = text.count("\n") + 1 if text else 0

        title = None
        for line in text.splitlines()[:20]:
            line = line.strip("# ").strip()
            if looks_like_title(line):
                title = line
                break
        out["title"] = title or out["title"]
        out["text"] = text[:5000]
        return out
    except Exception as e:
        out["extraction_status"] = f"error:{type(e).__name__}"
        return out


def parse_ris_text(text: str) -> Dict[str, Any]:
    data: Dict[str, Any] = {"authors": []}
    for raw in text.splitlines():
        line = raw.strip()
        if len(line) < 6 or line[2:6] != "  - ":
            continue
        tag = line[:2]
        value = line[6:].strip()
        if tag in {"TI", "T1"}:
            data["title"] = value
        elif tag in {"JO", "JF", "T2"}:
            data["journal"] = value
        elif tag in {"AU", "A1"}:
            data.setdefault("authors", []).append(value)
        elif tag in {"PY", "Y1"}:
            data["year"] = value[:4]
        elif tag == "DO":
            data["doi"] = value
    return data


def extract_ris(path: Path) -> Dict[str, Any]:
    out = {"text": "", "title": filename_to_title(path.name), "extraction_status": "ok", "authors": [], "journal": None, "year": None}
    try:
        text = path.read_text(encoding="utf-8", errors="ignore")
        parsed = parse_ris_text(text)
        out.update({k: v for k, v in parsed.items() if k in out or k in {"doi", "title"}})
        title = parsed.get("title") or out["title"]
        journal = parsed.get("journal")
        year = parsed.get("year")
        authors = parsed.get("authors", [])
        parts = [f"Citation export for {title}."]
        if journal or year:
            parts.append(f"Targets {journal or 'an unidentified venue'}{f' ({year})' if year else ''}.")
        if authors:
            parts.append(f"Contains references associated with {len(authors)} listed author entries.")
        out["text"] = " ".join(parts)
        out["title"] = title
        return out
    except Exception as e:
        out["extraction_status"] = f"error:{type(e).__name__}"
        return out


def extract_xlsx(path: Path) -> Dict[str, Any]:
    out: Dict[str, Any] = {
        "text": "",
        "title": filename_to_title(path.name),
        "sheet_names": [],
        "sheet_count": None,
        "sheet_headers": {},
        "extraction_status": "unavailable"
    }
    if load_workbook is None:
        return out

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheet_names = list(wb.sheetnames)
        out["sheet_names"] = sheet_names
        out["sheet_count"] = len(sheet_names)
        headers = {}
        summary_bits = []
        for sheet_name in sheet_names[:3]:
            ws = wb[sheet_name]
            first_rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
            header = []
            if first_rows:
                for cell in first_rows[0]:
                    if cell is None:
                        continue
                    val = str(cell).strip()
                    if val:
                        header.append(val[:80])
            headers[sheet_name] = header[:12]
            dims = f"{ws.max_row} rows x {ws.max_column} cols" if ws.max_row and ws.max_column else "size unknown"
            summary_bits.append(f"{sheet_name}: {dims}")
        out["sheet_headers"] = headers
        out["text"] = "Spreadsheet workbook. Sheets: " + "; ".join(summary_bits)
        out["extraction_status"] = "ok"
        return out
    except Exception as e:
        out["extraction_status"] = f"error:{type(e).__name__}"
        return out


def extract_image(path: Path) -> Dict[str, Any]:
    out: Dict[str, Any] = {
        "text": "",
        "title": filename_to_title(path.name),
        "dimensions": None,
        "extraction_status": "ok"
    }
    if Image is not None:
        try:
            with Image.open(path) as img:
                out["dimensions"] = {"width": img.width, "height": img.height}
        except Exception:
            pass
    return out


def extract_content(path: Path) -> Dict[str, Any]:
    ext = path.suffix.lower()
    if ext == ".pdf":
        return extract_pdf(path)
    if ext in {".md", ".markdown", ".txt", ".rst", ".tex", ".csv", ".json", ".yaml", ".yml", ".py", ".toml", ".ini", ".cfg"}:
        return extract_text_file(path)
    if ext == ".ris":
        return extract_ris(path)
    if ext == ".xlsx":
        return extract_xlsx(path)
    if ext in IMAGE_EXTS:
        return extract_image(path)
    return {"text": "", "title": filename_to_title(path.name), "extraction_status": "unsupported"}


def load_manifest(path: Optional[Path]) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, set]]:
    if not path:
        return {}, DEFAULT_EXCLUDES
    payload = json.loads(path.read_text(encoding="utf-8"))
    excludes = payload.get("excludes", {})
    exclude_dirs = set(excludes.get("dirs", [])) | set(DEFAULT_EXCLUDES["dirs"])
    exclude_file_names = set(excludes.get("file_names", [])) | set(DEFAULT_EXCLUDES["file_names"])
    exclude_suffixes = set(excludes.get("file_suffixes", [])) | set(DEFAULT_EXCLUDES["file_suffixes"])
    by_path: Dict[str, Dict[str, Any]] = {}
    for item in payload.get("files", []):
        rel = str(item["path"]).replace("\\", "/")
        by_path[rel] = item
    return by_path, {"dirs": exclude_dirs, "file_names": exclude_file_names, "file_suffixes": exclude_suffixes}


def should_skip(path: Path, exclude_rules: Dict[str, set], root: Path) -> bool:
    rel_parts = path.relative_to(root).parts
    if any(part in exclude_rules["dirs"] for part in rel_parts[:-1]):
        return True
    if path.name in exclude_rules["file_names"]:
        return True
    if path.suffix.lower() in exclude_rules["file_suffixes"]:
        return True
    return False


def collect_files(root: Path, manifest_lookup: Dict[str, Dict[str, Any]], exclude_rules: Dict[str, set]) -> List[FileRecord]:
    records: List[FileRecord] = []
    for abs_path in sorted(root.rglob("*")):
        if not abs_path.is_file():
            continue
        if should_skip(abs_path, exclude_rules, root):
            continue
        rel = safe_rel_path(abs_path, root)
        man = manifest_lookup.get(rel)
        sha = man.get("sha256") if man else None
        size = int(man.get("size")) if man and man.get("size") is not None else abs_path.stat().st_size
        if not sha:
            sha = sha256_file(abs_path)
        records.append(FileRecord(abs_path=abs_path, rel_path=rel, sha256=sha, size=size, modified_time_utc=file_mtime_iso(abs_path)))
    return records


def choose_primary_path(paths: Sequence[str]) -> str:
    return sorted(paths, key=lambda p: (p.count("/"), len(p), p.lower()))[0]


def build_routing_terms(title: str, summary: str, rel_paths: Sequence[str], topics: Sequence[str], doc_kind: str) -> List[str]:
    fields = [title, summary, " ".join(rel_paths), " ".join(topics), doc_kind]
    tokens = []
    for field in fields:
        tokens.extend(tokenize(field))
    counts = Counter(tokens)
    ranked = [tok for tok, _ in counts.most_common(40)]
    return ranked


def routing_text_for_entry(title: str, summary: str, rel_paths: Sequence[str], topics: Sequence[str], doc_kind: str) -> str:
    parts = [title, summary, " ".join(topics), doc_kind, " ".join(rel_paths)]
    text = " | ".join(p for p in parts if p)
    return compact(text, 1000)


def build_index(root: Path, output: Path, manifest_path: Optional[Path] = None, source_archive: Optional[str] = None) -> Dict[str, Any]:
    manifest_lookup, exclude_rules = load_manifest(manifest_path)
    files = collect_files(root, manifest_lookup, exclude_rules)

    grouped: Dict[str, List[FileRecord]] = defaultdict(list)
    for rec in files:
        grouped[rec.sha256].append(rec)

    entries: List[Dict[str, Any]] = []
    for sha, group in sorted(grouped.items(), key=lambda kv: choose_primary_path([g.rel_path for g in kv[1]]).lower()):
        rel_paths = sorted(g.rel_path for g in group)
        primary_rel = choose_primary_path(rel_paths)
        primary_abs = next(g.abs_path for g in group if g.rel_path == primary_rel)

        extracted = extract_content(primary_abs)
        title = extracted.get("title") or filename_to_title(primary_abs.name)
        snippet = extracted.get("text", "")
        topics = unique_keep_order(
            topic
            for rel in rel_paths
            for topic in infer_topics(rel)
        )
        doc_kind = infer_doc_kind(primary_abs.suffix.lower(), title, primary_rel, snippet)
        summary, summary_basis = summarize_findings(title, snippet, doc_kind)
        mime_type = guess_mime(primary_abs)

        metadata: Dict[str, Any] = {"representative_filename": primary_abs.name}
        for key in ("page_count", "line_count", "sheet_names", "sheet_count", "sheet_headers", "dimensions", "notes", "authors", "journal", "year", "doi"):
            value = extracted.get(key)
            if value not in (None, [], {}, ""):
                metadata[key] = value

        rel_paths_prefixed = [root.name + "/" + p for p in rel_paths]
        primary_rel_prefixed = root.name + "/" + primary_rel
        entry = {
            "id": sha[:12],
            "title": title,
            "summary": summary,
            "content_hash_sha256": sha,
            "primary_path": primary_rel_prefixed,
            "all_locations": rel_paths_prefixed,
            "location_count": len(rel_paths_prefixed),
            "topic_tags": topics,
            "file_type": primary_abs.suffix.lower().lstrip(".") or "unknown",
            "document_kind": doc_kind,
            "size_bytes": group[0].size,
            "duplicate_copies_bytes_total": sum(g.size for g in group),
            "modified_time_utc": max(g.modified_time_utc for g in group),
            "mime_type": mime_type,
            "summary_basis": summary_basis,
            "extraction_status": extracted.get("extraction_status", "unknown"),
            "metadata": metadata,
            "locate_hint": f"Same content appears in {len(rel_paths_prefixed)} location(s); primary copy listed first." if len(rel_paths_prefixed) > 1 else "Single known copy.",
            "routing_terms": build_routing_terms(title, summary, rel_paths_prefixed, topics, doc_kind),
            "routing_text": routing_text_for_entry(title, summary, rel_paths_prefixed, topics, doc_kind),
        }
        entries.append(entry)

    index = {
        "index_name": "external_research_index",
        "created_at_utc": utc_now_iso(),
        "root_folder": root.name,
        "source_archive": source_archive,
        "description": "Deduplicated research index for retrieval, grounding, and RAG routing.",
        "schema_notes": {
            "dedupe": "Entries are grouped by SHA-256 content hash. Duplicate file locations are preserved in all_locations.",
            "summaries": "Summaries are derived from abstracts/opening text where possible; otherwise filename/folder context is used.",
            "routing": "routing_terms and routing_text are intended for fast lexical routing before full retrieval."
        },
        "stats": {
            "physical_file_count": len(files),
            "unique_content_count": len(entries),
            "total_bytes_physical": sum(f.size for f in files),
            "dedupe_savings_bytes": max(0, sum(f.size for f in files) - sum(e["size_bytes"] for e in entries)),
        },
        "entries": entries
    }

    output.write_text(json.dumps(index, indent=2, ensure_ascii=False), encoding="utf-8")
    return index


def build_df(entries: Sequence[Dict[str, Any]]) -> Counter:
    df: Counter = Counter()
    for e in entries:
        doc_tokens = set(tokenize(" ".join([
            e.get("title", ""),
            e.get("summary", ""),
            " ".join(e.get("topic_tags", [])),
            e.get("primary_path", ""),
            e.get("document_kind", ""),
        ])))
        df.update(doc_tokens)
    return df


def score_entry(query: str, query_tokens: List[str], df: Counter, n_docs: int, entry: Dict[str, Any]) -> Tuple[float, Dict[str, Any]]:
    title = entry.get("title", "")
    summary = entry.get("summary", "")
    topics = " ".join(entry.get("topic_tags", []))
    path = entry.get("primary_path", "")
    doc_kind = entry.get("document_kind", "")

    field_weights = {
        "title": 5.0,
        "topic_tags": 4.0,
        "summary": 2.8,
        "path": 1.8,
        "document_kind": 1.2,
        "routing_terms": 1.6,
    }

    fields = {
        "title": tokenize(title),
        "topic_tags": tokenize(topics),
        "summary": tokenize(summary),
        "path": tokenize(path),
        "document_kind": tokenize(doc_kind),
        "routing_terms": list(entry.get("routing_terms", [])),
    }

    score = 0.0
    matches: Dict[str, List[str]] = defaultdict(list)
    for qt in query_tokens:
        idf = math.log((n_docs + 1) / (1 + df.get(qt, 0))) + 1.0
        for field_name, toks in fields.items():
            if qt in toks:
                score += field_weights[field_name] * idf
                matches[field_name].append(qt)

    query_low = query.lower().strip()
    joined_title = title.lower()
    joined_summary = summary.lower()
    joined_path = path.lower()

    if query_low and query_low in joined_title:
        score += 8.0
    if query_low and query_low in joined_summary:
        score += 5.0
    if query_low and query_low in joined_path:
        score += 4.0

    # Phrase bonuses for 2-grams.
    q_bigrams = [" ".join(query_tokens[i:i+2]) for i in range(max(0, len(query_tokens)-1))]
    for bg in q_bigrams:
        if bg and bg in joined_title:
            score += 3.0
        if bg and bg in joined_summary:
            score += 1.8
        if bg and bg in joined_path:
            score += 1.2

    explanation = {
        "matched_fields": {k: sorted(set(v)) for k, v in matches.items() if v},
        "title": title,
        "summary": summary,
        "primary_path": path,
        "topic_tags": entry.get("topic_tags", []),
        "document_kind": doc_kind,
    }
    return score, explanation


def route_query(index: Dict[str, Any], query: str, top_k: int = 8) -> Dict[str, Any]:
    entries = index.get("entries", [])
    query_tokens = tokenize(query)
    if not query_tokens:
        return {"query": query, "top_k": top_k, "results": []}

    df = build_df(entries)
    n_docs = max(1, len(entries))
    scored = []
    for e in entries:
        score, why = score_entry(query, query_tokens, df, n_docs, e)
        if score > 0:
            scored.append((score, e, why))

    scored.sort(key=lambda x: (-x[0], x[1].get("title", "").lower()))
    results = []
    for rank, (score, entry, why) in enumerate(scored[:top_k], start=1):
        results.append({
            "rank": rank,
            "score": round(score, 4),
            "id": entry["id"],
            "title": entry["title"],
            "summary": entry["summary"],
            "primary_path": entry["primary_path"],
            "all_locations": entry["all_locations"],
            "topic_tags": entry["topic_tags"],
            "document_kind": entry["document_kind"],
            "why": why["matched_fields"],
        })
    return {
        "query": query,
        "query_tokens": query_tokens,
        "top_k": top_k,
        "results": results,
    }


def cmd_build(args: argparse.Namespace) -> int:
    root = Path(args.root).expanduser().resolve()
    output = Path(args.output).expanduser().resolve()
    manifest = Path(args.manifest).expanduser().resolve() if args.manifest else None
    index = build_index(root=root, output=output, manifest_path=manifest, source_archive=args.source_archive)
    print(json.dumps({
        "ok": True,
        "output": str(output),
        "stats": index["stats"],
        "created_at_utc": index["created_at_utc"]
    }, indent=2))
    return 0


def cmd_route(args: argparse.Namespace) -> int:
    index_path = Path(args.index).expanduser().resolve()
    index = json.loads(index_path.read_text(encoding="utf-8"))
    result = route_query(index=index, query=args.query, top_k=args.top_k)
    print(json.dumps(result, indent=2 if args.pretty else None, ensure_ascii=False))
    return 0


def cmd_build_and_route(args: argparse.Namespace) -> int:
    root = Path(args.root).expanduser().resolve()
    output = Path(args.output).expanduser().resolve()
    manifest = Path(args.manifest).expanduser().resolve() if args.manifest else None
    index = build_index(root=root, output=output, manifest_path=manifest, source_archive=args.source_archive)
    result = route_query(index=index, query=args.query, top_k=args.top_k)
    print(json.dumps({
        "index_stats": index["stats"],
        "route_result": result,
        "output": str(output)
    }, indent=2, ensure_ascii=False))
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build a research index and route RAG queries.")
    sub = parser.add_subparsers(dest="command", required=True)

    p_build = sub.add_parser("build", help="Build the research index JSON.")
    p_build.add_argument("--root", required=True, help="Root research folder to scan.")
    p_build.add_argument("--output", required=True, help="Where to write the index JSON.")
    p_build.add_argument("--manifest", help="Optional manifest JSON with precomputed sha256/size and exclude rules.")
    p_build.add_argument("--source-archive", help="Optional name/path of the original source archive.")
    p_build.set_defaults(func=cmd_build)

    p_route = sub.add_parser("route", help="Route a query against an existing index.")
    p_route.add_argument("--index", required=True, help="Path to the built index JSON.")
    p_route.add_argument("--query", required=True, help="The user query to route.")
    p_route.add_argument("--top-k", type=int, default=8, help="Number of top candidates to return.")
    p_route.add_argument("--pretty", action="store_true", help="Pretty-print the JSON result.")
    p_route.set_defaults(func=cmd_route)

    p_both = sub.add_parser("build-and-route", help="Build an index and immediately route a query.")
    p_both.add_argument("--root", required=True, help="Root research folder to scan.")
    p_both.add_argument("--output", required=True, help="Where to write the index JSON.")
    p_both.add_argument("--manifest", help="Optional manifest JSON with precomputed sha256/size and exclude rules.")
    p_both.add_argument("--source-archive", help="Optional name/path of the original source archive.")
    p_both.add_argument("--query", required=True, help="The user query to route.")
    p_both.add_argument("--top-k", type=int, default=8, help="Number of top candidates to return.")
    p_both.set_defaults(func=cmd_build_and_route)

    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    return int(args.func(args))


if __name__ == "__main__":
    raise SystemExit(main())
